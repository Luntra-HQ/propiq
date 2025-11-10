"""
PropIQ Excel/CSV Exporter
Sprint 10-11: Features & Export

Export property analysis data to Excel and CSV formats for data portability.
"""

from typing import Dict, Any, List, Optional
from datetime import datetime
import csv
from io import StringIO, BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from config.logging_config import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """Export analysis data to Excel format"""

    def __init__(self):
        self.brand_color = '8B5CF6'  # PropIQ purple

    def export_analysis_to_excel(self, analysis_data: Dict[str, Any]) -> Optional[bytes]:
        """
        Export single analysis to Excel with multiple sheets

        Args:
            analysis_data: Property analysis dictionary

        Returns:
            Excel file as bytes, or None if export fails
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl not available. Cannot export to Excel.")
            return None

        try:
            wb = Workbook()

            # Sheet 1: Summary
            ws_summary = wb.active
            ws_summary.title = "Summary"
            self._create_summary_sheet(ws_summary, analysis_data)

            # Sheet 2: Financial Details
            ws_financial = wb.create_sheet("Financial Details")
            self._create_financial_sheet(ws_financial, analysis_data)

            # Sheet 3: Assumptions
            ws_assumptions = wb.create_sheet("Assumptions")
            self._create_assumptions_sheet(ws_assumptions, analysis_data)

            # Save to bytes
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            logger.info(f"Exported analysis {analysis_data.get('id')} to Excel")
            return excel_buffer.getvalue()

        except Exception as e:
            logger.error(f"Failed to export to Excel: {e}", exc_info=True)
            return None

    def _create_summary_sheet(self, ws, data: Dict[str, Any]):
        """Create summary sheet with key metrics"""

        # Header
        ws['A1'] = 'PropIQ Property Analysis Summary'
        ws['A1'].font = Font(size=16, bold=True, color=self.brand_color)

        # Property info
        ws['A3'] = 'Property Address:'
        ws['B3'] = data.get('address', 'N/A')
        ws['A4'] = 'City:'
        ws['B4'] = data.get('city', 'N/A')
        ws['A5'] = 'State:'
        ws['B5'] = data.get('state', 'N/A')

        # Deal Score
        ws['A7'] = 'Deal Score:'
        ws['B7'] = data.get('deal_score', 0)
        ws['B7'].font = Font(size=14, bold=True)

        # Key Metrics
        ws['A9'] = 'Key Metrics'
        ws['A9'].font = Font(bold=True)

        metrics = [
            ('Monthly Cash Flow:', data.get('monthly_cash_flow', 0)),
            ('Cash-on-Cash Return:', f"{data.get('cash_on_cash_return', 0):.2f}%"),
            ('Cap Rate:', f"{data.get('cap_rate', 0):.2f}%"),
            ('Annual ROI:', f"{data.get('roi', 0):.2f}%"),
        ]

        row = 10
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_financial_sheet(self, ws, data: Dict[str, Any]):
        """Create financial details sheet"""

        ws['A1'] = 'Financial Analysis'
        ws['A1'].font = Font(size=14, bold=True)

        # Income
        ws['A3'] = 'Income'
        ws['A3'].font = Font(bold=True)
        ws['A4'] = 'Monthly Rent'
        ws['B4'] = data.get('monthly_rent', 0)
        ws['A5'] = 'Annual Rent'
        ws['B5'] = data.get('monthly_rent', 0) * 12

        # Expenses
        ws['A7'] = 'Monthly Expenses'
        ws['A7'].font = Font(bold=True)

        expenses = [
            ('Mortgage (P&I)', data.get('monthly_mortgage', 0)),
            ('Property Taxes', data.get('monthly_property_tax', 0)),
            ('Insurance', data.get('monthly_insurance', 0)),
            ('Maintenance', data.get('monthly_maintenance', 0)),
            ('Vacancy Reserve', data.get('monthly_vacancy', 0)),
        ]

        row = 8
        total_expenses = 0
        for label, value in expenses:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            total_expenses += value
            row += 1

        ws[f'A{row}'] = 'Total Expenses'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_expenses
        ws[f'B{row}'].font = Font(bold=True)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _create_assumptions_sheet(self, ws, data: Dict[str, Any]):
        """Create assumptions sheet"""

        ws['A1'] = 'Analysis Assumptions'
        ws['A1'].font = Font(size=14, bold=True)

        assumptions = [
            ('Purchase Price', data.get('purchase_price', 0)),
            ('Down Payment %', data.get('down_payment_percent', 0)),
            ('Interest Rate %', data.get('interest_rate', 0)),
            ('Loan Term (years)', data.get('loan_term_years', 0)),
            ('Property Tax Rate %', data.get('property_tax_rate', 0)),
        ]

        row = 3
        for label, value in assumptions:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20


class CSVExporter:
    """Export analysis data to CSV format"""

    def export_analysis_to_csv(self, analysis_data: Dict[str, Any]) -> str:
        """
        Export analysis to CSV format

        Args:
            analysis_data: Property analysis dictionary

        Returns:
            CSV as string
        """
        output = StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(['PropIQ Property Analysis Export'])
        writer.writerow([])

        # Property Info
        writer.writerow(['Property Information'])
        writer.writerow(['Address', analysis_data.get('address', 'N/A')])
        writer.writerow(['City', analysis_data.get('city', 'N/A')])
        writer.writerow(['State', analysis_data.get('state', 'N/A')])
        writer.writerow([])

        # Key Metrics
        writer.writerow(['Key Metrics'])
        writer.writerow(['Deal Score', analysis_data.get('deal_score', 0)])
        writer.writerow(['Monthly Cash Flow', analysis_data.get('monthly_cash_flow', 0)])
        writer.writerow(['Cash-on-Cash Return %', analysis_data.get('cash_on_cash_return', 0)])
        writer.writerow(['Cap Rate %', analysis_data.get('cap_rate', 0)])
        writer.writerow(['Annual ROI %', analysis_data.get('roi', 0)])
        writer.writerow([])

        # Financial Details
        writer.writerow(['Financial Details'])
        writer.writerow(['Monthly Rent', analysis_data.get('monthly_rent', 0)])
        writer.writerow(['Monthly Mortgage', analysis_data.get('monthly_mortgage', 0)])
        writer.writerow(['Monthly Property Tax', analysis_data.get('monthly_property_tax', 0)])
        writer.writerow(['Monthly Insurance', analysis_data.get('monthly_insurance', 0)])
        writer.writerow(['Monthly Maintenance', analysis_data.get('monthly_maintenance', 0)])

        logger.info(f"Exported analysis {analysis_data.get('id')} to CSV")
        return output.getvalue()

    def export_multiple_analyses_to_csv(self, analyses: List[Dict[str, Any]]) -> str:
        """
        Export multiple analyses to CSV for comparison

        Args:
            analyses: List of analysis dictionaries

        Returns:
            CSV as string with comparison table
        """
        output = StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(['Property Comparison Export - PropIQ'])
        writer.writerow([])

        # Column headers
        headers = ['Metric'] + [a.get('address', f"Property {i+1}") for i, a in enumerate(analyses)]
        writer.writerow(headers)

        # Metrics rows
        metrics = [
            ('Deal Score', 'deal_score'),
            ('Purchase Price', 'purchase_price'),
            ('Monthly Cash Flow', 'monthly_cash_flow'),
            ('Cash-on-Cash Return %', 'cash_on_cash_return'),
            ('Cap Rate %', 'cap_rate'),
            ('Annual ROI %', 'roi'),
        ]

        for label, key in metrics:
            row = [label] + [a.get(key, 'N/A') for a in analyses]
            writer.writerow(row)

        logger.info(f"Exported {len(analyses)} analyses to comparison CSV")
        return output.getvalue()


# Global instances
excel_exporter = ExcelExporter()
csv_exporter = CSVExporter()


def export_to_excel(analysis_data: Dict[str, Any]) -> Optional[bytes]:
    """Convenience function to export to Excel"""
    return excel_exporter.export_analysis_to_excel(analysis_data)


def export_to_csv(analysis_data: Dict[str, Any]) -> str:
    """Convenience function to export to CSV"""
    return csv_exporter.export_analysis_to_csv(analysis_data)


def export_comparison_to_csv(analyses: List[Dict[str, Any]]) -> str:
    """Convenience function to export comparison to CSV"""
    return csv_exporter.export_multiple_analyses_to_csv(analyses)
